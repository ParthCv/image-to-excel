"""
main.py  --  Freezer box grid photo  ->  Excel (.xlsx)

One file. No modules to wire up. Run it, point it at a photo, get a spreadsheet.

    python main.py photo.jpg
    python main.py photo.jpg -o AGM_boxA.xlsx

WHAT IT DOES
    1. Finds the ruled 10x10 grid in a phone photo  (OpenCV, no ML, deterministic)
    2. Warps it flat, splits it into 100 cells
    3. Reads each cell locally with EasyOCR       (the ONLY model; image never leaves the machine)
    4. Writes an .xlsx that matches example.xlsx exactly, with suspicious cells
       highlighted amber + a comment saying why, so the human review step inside
       Excel knows exactly where to look.

DESIGN RULE (do not remove): nothing is ever auto-corrected. Every check only
raises a flag. A silently "fixed" MB-CR -> MB-CP is an invisible permanent error.
The machine narrows attention; the human decides.

PRIVACY: the photo is patient data. It is read entirely on this machine. There is
no network call with the image anywhere in this file. After the first run (which
downloads the EasyOCR weights once), you may set HF_HUB_OFFLINE=1 / no internet and
it still works -- proof, not a promise.
"""

import os
import re
import sys
import argparse

import numpy as np
import cv2

EXPECTED_ROWS = 10          
EXPECTED_COLS = 10          
GRID_TOLERANCE = 1          
MAX_SPACING_CV = 0.15       
WORK_RES = 1600             

ID_RE = re.compile(r"^[A-Z]{3}-\d{1,4}$")          
TYPES = {"DNA", "RNA"}
MODIFIER_RE = re.compile(r"^[A-Z]{2,3}(-[A-Z]{2})?$")  

BLANK_TOKEN = "(blank)"     

INK_BLANK_MAX = 0.045
INK_WRITTEN_MIN = 0.065

OCR_ALLOWLIST = "ABCDEFGHIJKLMNOPQRSTUVWXYZ0123456789-. "

OCR_CONF_FLAG = 0.45        

def _downscale(gray):
    h, w = gray.shape
    s = WORK_RES / max(h, w)
    if s < 1.0:
        return cv2.resize(gray, (int(w * s), int(h * s)), interpolation=cv2.INTER_AREA), s
    return gray, 1.0

def _order_pts(p):
    r = np.zeros((4, 2), dtype="float32")
    s = p.sum(1)
    d = np.diff(p, axis=1)
    r[0] = p[np.argmin(s)]   
    r[2] = p[np.argmax(s)]   
    r[1] = p[np.argmin(d)]   
    r[3] = p[np.argmax(d)]   
    return r

def _line_mask(gray, w, h):
    thr = cv2.adaptiveThreshold(gray, 255, cv2.ADAPTIVE_THRESH_MEAN_C,
                                cv2.THRESH_BINARY_INV, 25, 10)
    hk = cv2.getStructuringElement(cv2.MORPH_RECT, (max(15, w // 40), 1))
    vk = cv2.getStructuringElement(cv2.MORPH_RECT, (1, max(15, h // 40)))
    horiz = cv2.dilate(cv2.erode(thr, hk), hk)
    vert = cv2.dilate(cv2.erode(thr, vk), vk)
    return horiz | vert

def _find_table_quad(gray):
    h, w = gray.shape
    lines = cv2.dilate(_line_mask(gray, w, h), np.ones((3, 3), np.uint8), iterations=2)
    cnts, _ = cv2.findContours(lines, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    if not cnts:
        return None
    c = max(cnts, key=cv2.contourArea)
    if cv2.contourArea(c) < 0.12 * h * w:
        return None
    approx = cv2.approxPolyDP(c, 0.02 * cv2.arcLength(c, True), True)
    if len(approx) == 4:
        return approx.reshape(4, 2).astype("float32")
    return cv2.boxPoints(cv2.minAreaRect(c)).astype("float32")

def _warp(gray, quad):
    r = _order_pts(quad)
    (tl, tr, br, bl) = r
    W = int(max(np.linalg.norm(br - bl), np.linalg.norm(tr - tl)))
    H = int(max(np.linalg.norm(tr - br), np.linalg.norm(tl - bl)))
    M = cv2.getPerspectiveTransform(
        r, np.array([[0, 0], [W - 1, 0], [W - 1, H - 1], [0, H - 1]], dtype="float32"))
    return cv2.warpPerspective(gray, M, (W, H))

def _tight_crop(warped):
    h, w = warped.shape
    thr = cv2.adaptiveThreshold(warped, 255, cv2.ADAPTIVE_THRESH_MEAN_C,
                                cv2.THRESH_BINARY_INV, 25, 10)
    hk = cv2.getStructuringElement(cv2.MORPH_RECT, (w // 3, 1))
    vk = cv2.getStructuringElement(cv2.MORPH_RECT, (1, h // 3))
    hm = cv2.dilate(cv2.erode(thr, hk), hk)
    vm = cv2.dilate(cv2.erode(thr, vk), vk)
    ca = vm.sum(0).astype(float)
    ra = hm.sum(1).astype(float)
    cx = np.where(ca > 0.15 * ca.max())[0]
    ry = np.where(ra > 0.15 * ra.max())[0]
    if len(cx) < 2 or len(ry) < 2:
        return warped
    return warped[ry[0]:ry[-1] + 1, cx[0]:cx[-1] + 1]

def _peaks(proj, expected, span):
    proj = proj.astype(float)
    proj = proj / (proj.max() + 1e-9)
    minsep = int(span / (expected + 3))
    order = sorted([i for i in range(len(proj)) if proj[i] > 0.15], key=lambda i: -proj[i])
    keep = []
    for i in order:
        if all(abs(i - k) >= minsep for k in keep):
            keep.append(i)
    return sorted(keep)

def _grid_lines(warped):
    h, w = warped.shape
    thr = cv2.adaptiveThreshold(warped, 255, cv2.ADAPTIVE_THRESH_MEAN_C,
                                cv2.THRESH_BINARY_INV, 25, 10)
    hk = cv2.getStructuringElement(cv2.MORPH_RECT, (w // 3, 1))
    vk = cv2.getStructuringElement(cv2.MORPH_RECT, (1, h // 3))
    hm = cv2.dilate(cv2.erode(thr, hk), hk)
    vm = cv2.dilate(cv2.erode(thr, vk), vk)
    rows = _peaks(hm.sum(1), EXPECTED_ROWS, h)
    cols = _peaks(vm.sum(0), EXPECTED_COLS, w)
    return rows, cols

def _spacing_cv(lines):
    if len(lines) < 3:
        return 9.9
    d = np.diff(lines)
    return float(np.std(d) / (np.mean(d) + 1e-9))

class GridError(Exception):
    pass

def detect_grid(bgr):
    gray = cv2.cvtColor(bgr, cv2.COLOR_BGR2GRAY)
    small, _ = _downscale(gray)
    quad = _find_table_quad(small)
    if quad is None:
        raise GridError("No ruled grid found. Is the whole box form in frame, "
                        "reasonably flat and well lit?")
    warped = _tight_crop(_warp(small, quad))
    rows, cols = _grid_lines(warped)
    nrows, ncols = max(0, len(rows) - 1), max(0, len(cols) - 1)
    rcv, ccv = _spacing_cv(rows), _spacing_cv(cols)

    lo_r, hi_r = EXPECTED_ROWS - GRID_TOLERANCE, EXPECTED_ROWS + GRID_TOLERANCE
    lo_c, hi_c = EXPECTED_COLS - GRID_TOLERANCE, EXPECTED_COLS + GRID_TOLERANCE
    if not (lo_r <= nrows <= hi_r and lo_c <= ncols <= hi_c):
        raise GridError(
            "Grid read as {}x{} but these forms are always {}x{}.\n"
            "        This is almost always a photo problem, not a real small grid.\n"
            "        Retake it: shoot at a slight ANGLE (not straight down), fill the\n"
            "        frame with the form, avoid glare/washout on any edge."
            .format(nrows, ncols, EXPECTED_ROWS, EXPECTED_COLS))
    if rcv > MAX_SPACING_CV or ccv > MAX_SPACING_CV:
        raise GridError(
            "Grid lines are too irregular (row cv={:.3f}, col cv={:.3f}) to trust.\n"
            "        Retake the photo flatter / less angled.".format(rcv, ccv))

    rows = _uniform(rows, EXPECTED_ROWS + 1)
    cols = _uniform(cols, EXPECTED_COLS + 1)
    return warped, rows, cols

def _uniform(lines, n):
    if len(lines) == n:
        return lines
    lo, hi = lines[0], lines[-1]
    return [int(round(lo + (hi - lo) * i / (n - 1))) for i in range(n)]

def _ink_fraction(cell):
    thr = cv2.adaptiveThreshold(cell, 255, cv2.ADAPTIVE_THRESH_MEAN_C,
                                cv2.THRESH_BINARY_INV, 15, 8)
    h, w = thr.shape
    m = int(0.12 * min(h, w))  
    inner = thr[m:h - m, m:w - m] if h > 2 * m and w > 2 * m else thr
    return float((inner > 0).mean())

def extract_cells(warped, rows, cols):
    grid = []
    for ri in range(len(rows) - 1):
        line = []
        for ci in range(len(cols) - 1):
            y0, y1 = rows[ri], rows[ri + 1]
            x0, x1 = cols[ci], cols[ci + 1]
            pad_y = int(0.06 * (y1 - y0))
            pad_x = int(0.06 * (x1 - x0))
            crop = warped[y0 + pad_y:y1 - pad_y, x0 + pad_x:x1 - pad_x]
            if crop.size == 0:
                line.append(dict(crop=None, ink=0.0, state="blank"))
                continue
            ink = _ink_fraction(crop)
            if ink <= INK_BLANK_MAX:
                state = "blank"
            elif ink >= INK_WRITTEN_MIN:
                state = "written"
            else:
                state = "uncertain"
            line.append(dict(crop=crop, ink=ink, state=state))
        grid.append(line)
    return grid

class Recognizer:
    def __init__(self, use_gpu=None):
        import easyocr  
        if use_gpu is None:
            use_gpu = _cuda_available()
        self.reader = easyocr.Reader(["en"], gpu=use_gpu, verbose=False)

    def read_lines(self, crop):
        if crop is None or crop.size == 0:
            return []
        up = cv2.resize(crop, None, fx=2.0, fy=2.0, interpolation=cv2.INTER_CUBIC)
        results = self.reader.readtext(
            up, detail=1, paragraph=False, allowlist=OCR_ALLOWLIST)
        results.sort(key=lambda r: min(p[1] for p in r[0]))
        out = []
        for box, text, conf in results:
            t = text.strip().upper().replace(" ", "").replace("O0", "O")
            if t:
                out.append((t, float(conf)))
        return out

def _cuda_available():
    try:
        import torch
        return bool(torch.cuda.is_available())
    except Exception:
        return False

class Cell:
    __slots__ = ("id", "modifier", "type", "raw", "flags", "crop", "row", "col")

    def __init__(self, crop=None, row=0, col=0):
        self.id = ""
        self.modifier = ""
        self.type = ""
        self.raw = []
        self.flags = []
        self.crop = crop
        self.row = row
        self.col = col

    @property
    def is_blank(self):
        return not (self.id or self.modifier or self.type or self.raw)

    def value(self):
        if self.is_blank:
            return BLANK_TOKEN
        idline = self.id if self.id else "?"
        if self.modifier:
            idline = idline + " " + self.modifier
        typ = self.type if self.type else ""
        return idline + ("\n" + typ if typ else "")

def _clean_id(tok):
    t = tok.replace(" ", "").upper()
    return t

def assemble_cell(lines):
    cell = Cell()
    cell.raw = list(lines)
    if not lines:
        return cell

    texts = [t for t, _ in lines]
    confs = [c for _, c in lines]

    type_idx = None
    for i, t in enumerate(texts):
        if t in TYPES:
            type_idx = i
            break
    if type_idx is not None:
        cell.type = texts[type_idx]
    else:
        cell.flags.append("no clear DNA/RNA line")

    id_line = None
    for i, t in enumerate(texts):
        if i == type_idx:
            continue
        id_line = t
        break

    if id_line is not None:
        m = re.match(r"^([A-Z]{3}-\d{1,4})(.*)$", id_line)
        if m:
            cell.id = m.group(1)
            rest = m.group(2).strip(" -")
            if rest:
                cell.modifier = rest
        else:
            cell.id = _clean_id(id_line)

    if cell.id and not ID_RE.match(cell.id):
        cell.flags.append("ID '{}' doesn't match AAA-#### pattern".format(cell.id))
    if cell.modifier and not MODIFIER_RE.match(cell.modifier):
        cell.flags.append("unexpected modifier '{}'".format(cell.modifier))
    if type_idx is not None and any(t not in TYPES and not ID_RE.match(t) and
                                    t != cell.modifier for t in texts):
        leftover = [t for t in texts if t not in TYPES and t != cell.id
                    and t != cell.modifier]
        if leftover:
            cell.flags.append("unrecognised text: " + " ".join(leftover))
    if confs and min(confs) < OCR_CONF_FLAG:
        cell.flags.append("low OCR confidence ({:.2f})".format(min(confs)))
    return cell

def vocab_flags(grid_cells):
    from collections import Counter
    ids = [c.id for row in grid_cells for c in row if c.id and ID_RE.match(c.id)]
    counts = Counter(ids)
    common = {k for k, v in counts.items() if v >= 2}
    for row in grid_cells:
        for c in row:
            if c.id and counts.get(c.id, 0) == 1 and c.id not in common:
                near = [k for k in common if _lev(c.id, k) == 1]
                if near:
                    c.flags.append("ID '{}' is 1 edit from '{}' (seen {}x) - misread?"
                                   .format(c.id, near[0], counts[near[0]]))
    return grid_cells

def _lev(a, b):
    if a == b:
        return 0
    la, lb = len(a), len(b)
    if abs(la - lb) > 1:
        return 2
    prev = list(range(lb + 1))
    for i, ca in enumerate(a, 1):
        cur = [i]
        for j, cb in enumerate(b, 1):
            cur.append(min(prev[j] + 1, cur[-1] + 1, prev[j - 1] + (ca != cb)))
        prev = cur
    return prev[-1]

def _read_grid(bgr, recognizer, on_progress=None):
    warped, rows, cols = detect_grid(bgr)
    cells_meta = extract_cells(warped, rows, cols)
    out = []
    total = sum(1 for r in cells_meta for c in r if c["state"] != "blank")
    done = 0
    for ri, row in enumerate(cells_meta):
        line = []
        for ci, meta in enumerate(row):
            if meta["state"] == "blank":
                cell = Cell(crop=meta["crop"], row=ri, col=ci)
                line.append(cell)
                continue
            lines = recognizer.read_lines(meta["crop"])
            cell = assemble_cell(lines)
            cell.crop = meta["crop"]
            cell.row = ri
            cell.col = ci
            if meta["state"] == "uncertain":
                cell.flags.append("ink level ambiguous (blank or faint?) -- verify")
            line.append(cell)
            done += 1
            if on_progress:
                on_progress(done, total)
        out.append(line)
    return out

def _regex_hits(grid):
    return sum(1 for row in grid for c in row if c.id and ID_RE.match(c.id))

def process_image(path, recognizer, on_progress=None):
    bgr = cv2.imread(path)
    if bgr is None:
        raise GridError("Could not open image: {}".format(path))

    grid = _read_grid(bgr, recognizer, on_progress)
    hits = _regex_hits(grid)

    if hits < 0.6 * EXPECTED_ROWS * EXPECTED_COLS:
        flipped = cv2.rotate(bgr, cv2.ROTATE_180)
        try:
            grid2 = _read_grid(flipped, recognizer, on_progress)
            if _regex_hits(grid2) > hits:
                grid = grid2
        except GridError:
            pass

    return vocab_flags(grid)

def _a1(r, c):
    return "{}{}".format(chr(65 + c), r + 1)

def review_and_edit(grid):
    import tkinter as tk
    from tkinter import ttk
    from PIL import Image, ImageTk
    from collections import Counter

    cells = [c for row in grid for c in row if not c.is_blank]
    if not cells: 
        return False

    counts = Counter([c.id for c in cells if c.id])
    common = [k for k, v in counts.items() if v >= 2]

    root = tk.Tk()
    root.title("Review and Edit Grid")
    root.geometry("900x700")

    canvas = tk.Canvas(root)
    scroll = ttk.Scrollbar(root, orient="vertical", command=canvas.yview)
    frame = ttk.Frame(canvas)
    
    frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
    canvas.create_window((0, 0), window=frame, anchor="nw")
    canvas.configure(yscrollcommand=scroll.set)
    
    canvas.pack(side="left", fill="both", expand=True)
    scroll.pack(side="right", fill="y")

    tk_imgs = []
    vars_list = []

    for c in cells:
        row_bg = "#FFF2CC" if c.flags else None
        row_frame = tk.Frame(frame, bd=1, relief="solid", bg=row_bg)
        row_frame.pack(fill="x", padx=5, pady=2)
        
        if c.crop is not None:
            img = Image.fromarray(c.crop).resize((120, 40))
            tk_img = ImageTk.PhotoImage(img)
            tk_imgs.append(tk_img)
            lbl = tk.Label(row_frame, image=tk_img, bg=row_bg)
            lbl.pack(side="left", padx=5)
        
        ref_lbl = tk.Label(row_frame, text=_a1(c.row, c.col), width=4, bg=row_bg)
        ref_lbl.pack(side="left")

        v_id = tk.StringVar(value=c.id)
        cb_id = ttk.Combobox(row_frame, textvariable=v_id, values=common, width=12)
        cb_id.pack(side="left", padx=5)

        v_mod = tk.StringVar(value=c.modifier)
        ent_mod = tk.Entry(row_frame, textvariable=v_mod, width=10)
        ent_mod.pack(side="left", padx=5)

        v_type = tk.StringVar(value=c.type)
        cb_type = ttk.Combobox(row_frame, textvariable=v_type, values=["", "DNA", "RNA"], width=5, state="readonly")
        cb_type.pack(side="left", padx=5)

        flags_txt = " | ".join(c.flags)
        f_lbl = tk.Label(row_frame, text=flags_txt, fg="red", bg=row_bg)
        f_lbl.pack(side="left", padx=10)

        vars_list.append((c, v_id, v_mod, v_type))

    changed = False
    def on_close():
        nonlocal changed
        for c, v_id, v_mod, v_type in vars_list:
            nid, nmod, ntype = v_id.get().strip(), v_mod.get().strip(), v_type.get().strip()
            if nid != c.id or nmod != c.modifier or ntype != c.type:
                c.id, c.modifier, c.type = nid, nmod, ntype
                c.flags = []
                changed = True
        root.destroy()

    btn = tk.Button(root, text="Save & Close", command=on_close, bg="green", fg="white")
    btn.pack(side="bottom", fill="x", pady=5)
    root.protocol("WM_DELETE_WINDOW", on_close)
    
    root.mainloop()
    return changed

def write_xlsx(grid, out_path, source_name=""):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.comments import Comment

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    font = Font(name="Arial", bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    amber = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

    n_flagged = 0
    for ri, row in enumerate(grid, start=1):
        for ci, cell in enumerate(row, start=1):
            xc = ws.cell(row=ri, column=ci, value=cell.value())
            xc.font = font
            xc.alignment = center
            if cell.flags:
                xc.fill = amber
                xc.comment = Comment("FLAGGED - verify:\n- " +
                                     "\n- ".join(cell.flags), "freezer_grid")
                n_flagged += 1

    for ci in range(1, EXPECTED_COLS + 1):
        ws.column_dimensions[chr(64 + ci)].width = 14
    for ri in range(1, EXPECTED_ROWS + 1):
        ws.row_dimensions[ri].height = 34

    wb.save(out_path)
    return n_flagged

def _default_out(path):
    base = os.path.splitext(os.path.basename(path))[0]
    return os.path.join(os.path.dirname(os.path.abspath(path)), base + ".xlsx")

def main(argv=None):
    ap = argparse.ArgumentParser(
        description="Freezer box grid photo -> Excel. Reads locally; image never leaves the machine.")
    ap.add_argument("image", help="path to the phone photo of the box form")
    ap.add_argument("-o", "--out", help="output .xlsx path (default: alongside the image)")
    ap.add_argument("--cpu", action="store_true", help="force CPU even if a GPU is present")
    ap.add_argument("--review", action="store_true", help="open UI to review and correct reads")
    args = ap.parse_args(argv)

    if not os.path.exists(args.image):
        print("ERROR: file not found:", args.image)
        return 2

    out_path = args.out or _default_out(args.image)

    print("Reading grid from:", args.image)
    print("Loading the local text recogniser (first run downloads weights once)...")
    try:
        rec = Recognizer(use_gpu=False if args.cpu else None)
    except ImportError:
        print("\nERROR: EasyOCR isn't installed. Run:  pip install -r requirements.txt")
        return 3

    def progress(done, total):
        pct = int(100 * done / max(total, 1))
        sys.stdout.write("\r  reading cells... {:3d}%  ({}/{})".format(pct, done, total))
        sys.stdout.flush()

    try:
        grid = process_image(args.image, rec, on_progress=progress)
    except GridError as e:
        print("\n\nCANNOT PROCESS THIS PHOTO:\n        " + str(e))
        print("\n        (Nothing was written. This is deliberate: a wrong grid is worse\n"
              "        than no grid.)")
        return 4
    print()  

    if args.review and review_and_edit(grid):
        grid = vocab_flags(grid)

    n_flagged = write_xlsx(grid, out_path, source_name=os.path.basename(args.image))
    n_blank = sum(1 for row in grid for c in row if c.is_blank)
    n_filled = EXPECTED_ROWS * EXPECTED_COLS - n_blank

    print("\nDone.")
    print("  Wrote: {}".format(out_path))
    print("  {} filled cells, {} blank.".format(n_filled, n_blank))
    if n_flagged:
        print("  {} cell(s) flagged AMBER for review -- open the file and check the".format(n_flagged))
        print("  highlighted cells (hover for the reason) before uploading to Drive.")
    else:
        print("  No cells flagged. Still worth a quick eyeball before uploading.")
    return 0

if __name__ == "__main__":
    raise SystemExit(main())